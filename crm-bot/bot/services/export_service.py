import io

from openpyxl import Workbook
from openpyxl.styles import Font

from bot.database.models import Contest, Visitor


def build_participants_excel(contest: Contest, leaderboard: list[tuple[Visitor, int | None]]) -> bytes:
    """Konkurs ishtirokchilari ro'yxatini Excel fayl ko'rinishida (bytes)
    qaytaradi. Referal konkursida 'Referal soni' ustuniga son, RANDOM
    konkursida esa '-' yoziladi - LEKIN jadval tartibi (ustunlar) ikkalasida
    ham AYNAN BIR XIL bo'ladi."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ishtirokchilar"

    headers = ["№", "Ism-familiya", "Username", "Telegram ID", "Referal soni"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)

    for i, (visitor, count) in enumerate(leaderboard, start=1):
        ws.append([
            i,
            visitor.full_name,
            f"@{visitor.username}" if visitor.username else "-",
            visitor.telegram_id,
            count if count is not None else "-",
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
